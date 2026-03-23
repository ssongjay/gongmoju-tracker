from fastapi import APIRouter, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional
from io import BytesIO

from database import get_db
from scraper import fetch_ipo_list

router = APIRouter()


# --- Models ---

class MemberCreate(BaseModel):
    name: str


class IPORecordCreate(BaseModel):
    member_id: int
    stock_name: str
    broker: str
    ipo_price: int
    allocated: bool
    allocated_qty: int = 0
    sell_price: Optional[int] = None
    sell_date: Optional[str] = None
    memo: Optional[str] = None


class IPORecordUpdate(BaseModel):
    stock_name: Optional[str] = None
    broker: Optional[str] = None
    ipo_price: Optional[int] = None
    allocated: Optional[bool] = None
    allocated_qty: Optional[int] = None
    sell_price: Optional[int] = None
    sell_date: Optional[str] = None
    memo: Optional[str] = None


# --- Members ---

@router.get("/members")
def list_members():
    db = get_db()
    rows = db.execute("SELECT * FROM members ORDER BY name").fetchall()
    db.close()
    return [dict(r) for r in rows]


@router.post("/members")
def create_member(data: MemberCreate):
    db = get_db()
    cur = db.execute("INSERT INTO members (name) VALUES (?)", (data.name,))
    db.commit()
    member_id = cur.lastrowid
    db.close()
    return {"id": member_id, "name": data.name}


@router.delete("/members/{member_id}")
def delete_member(member_id: int):
    db = get_db()
    db.execute("DELETE FROM ipo_records WHERE member_id = ?", (member_id,))
    db.execute("DELETE FROM members WHERE id = ?", (member_id,))
    db.commit()
    db.close()
    return {"ok": True}


# --- IPO Records ---

@router.get("/records")
def list_records(member_id: Optional[int] = Query(None)):
    db = get_db()
    if member_id:
        rows = db.execute(
            "SELECT r.*, m.name as member_name FROM ipo_records r "
            "JOIN members m ON r.member_id = m.id "
            "WHERE r.member_id = ? ORDER BY r.created_at DESC",
            (member_id,),
        ).fetchall()
    else:
        rows = db.execute(
            "SELECT r.*, m.name as member_name FROM ipo_records r "
            "JOIN members m ON r.member_id = m.id "
            "ORDER BY r.created_at DESC"
        ).fetchall()
    db.close()
    return [dict(r) for r in rows]


@router.post("/records")
def create_record(data: IPORecordCreate):
    db = get_db()
    cur = db.execute(
        "INSERT INTO ipo_records (member_id, stock_name, broker, ipo_price, allocated, allocated_qty, sell_price, sell_date, memo) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
        (
            data.member_id,
            data.stock_name,
            data.broker,
            data.ipo_price,
            1 if data.allocated else 0,
            data.allocated_qty,
            data.sell_price,
            data.sell_date,
            data.memo,
        ),
    )
    db.commit()
    record_id = cur.lastrowid
    db.close()
    return {"id": record_id}


@router.put("/records/{record_id}")
def update_record(record_id: int, data: IPORecordUpdate):
    db = get_db()
    fields = []
    values = []
    for field, value in data.model_dump(exclude_unset=True).items():
        if field == "allocated" and value is not None:
            value = 1 if value else 0
        fields.append(f"{field} = ?")
        values.append(value)
    if fields:
        values.append(record_id)
        db.execute(
            f"UPDATE ipo_records SET {', '.join(fields)} WHERE id = ?", values
        )
        db.commit()
    db.close()
    return {"ok": True}


@router.delete("/records/{record_id}")
def delete_record(record_id: int):
    db = get_db()
    db.execute("DELETE FROM ipo_records WHERE id = ?", (record_id,))
    db.commit()
    db.close()
    return {"ok": True}


# --- Stats ---

@router.get("/stats")
def get_stats(member_id: Optional[int] = Query(None)):
    db = get_db()
    where = "WHERE r.member_id = ?" if member_id else ""
    params = (member_id,) if member_id else ()

    rows = db.execute(
        f"SELECT r.*, m.name as member_name FROM ipo_records r "
        f"JOIN members m ON r.member_id = m.id {where} "
        f"ORDER BY r.created_at DESC",
        params,
    ).fetchall()
    db.close()

    total_applied = len(rows)
    total_allocated = sum(1 for r in rows if r["allocated"])
    total_profit = 0
    total_invested = 0

    for r in rows:
        if r["allocated"] and r["sell_price"] and r["allocated_qty"] > 0:
            profit = (r["sell_price"] - r["ipo_price"]) * r["allocated_qty"]
            total_profit += profit
            total_invested += r["ipo_price"] * r["allocated_qty"]

    return {
        "total_applied": total_applied,
        "total_allocated": total_allocated,
        "allocation_rate": round(total_allocated / total_applied * 100, 1) if total_applied > 0 else 0,
        "total_profit": total_profit,
        "total_invested": total_invested,
        "profit_rate": round(total_profit / total_invested * 100, 1) if total_invested > 0 else 0,
    }


# --- IPO List (from 38.co.kr) ---

@router.get("/ipo-list")
def get_ipo_list(page: int = Query(1, ge=1)):
    try:
        return fetch_ipo_list(page=page)
    except Exception:
        return []


# --- Excel Export ---

@router.get("/export")
def export_excel(member_id: Optional[int] = Query(None)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    db = get_db()
    where = "WHERE r.member_id = ?" if member_id else ""
    params = (member_id,) if member_id else ()
    rows = db.execute(
        f"SELECT r.*, m.name as member_name FROM ipo_records r "
        f"JOIN members m ON r.member_id = m.id {where} "
        f"ORDER BY r.created_at DESC",
        params,
    ).fetchall()
    db.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "공모주 기록"

    headers = ["이름", "종목명", "증권사", "공모가", "배정여부", "배정수량", "매도가", "매도일", "수익금", "수익률", "메모"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row_idx, r in enumerate(rows, 2):
        profit = (r["sell_price"] - r["ipo_price"]) * r["allocated_qty"] if r["allocated"] and r["sell_price"] and r["allocated_qty"] > 0 else None
        profit_rate = round(profit / (r["ipo_price"] * r["allocated_qty"]) * 100, 1) if profit is not None and r["ipo_price"] > 0 else None

        values = [
            r["member_name"],
            r["stock_name"],
            r["broker"],
            r["ipo_price"],
            "O" if r["allocated"] else "X",
            r["allocated_qty"] if r["allocated"] else 0,
            r["sell_price"] if r["sell_price"] else "",
            r["sell_date"] if r["sell_date"] else "",
            f"{profit:,}원" if profit is not None else "",
            f"{profit_rate}%" if profit_rate is not None else "",
            r["memo"] if r["memo"] else "",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 14

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=gongmoju_records.xlsx"},
    )
